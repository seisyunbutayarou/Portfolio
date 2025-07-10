import os
import time
import random
import pandas as pd
import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from multiprocessing import Pool, cpu_count
from openpyxl import load_workbook

# --- 設定 ---
CHROME_DRIVER_PATH = r"" #クロームドライバーパス
AMAZON_LOGIN_URL = "https://www.amazon.co.jp/ap/signin???" #アマゾンログインURL
AMAZON_EMAIL = "" #メールアドレス
AMAZON_PASSWORD = "" #パスワード
INPUT_EXCEL = "input.xlsx"
OUTPUT_EXCEL = "output.xlsx"
NUM_WORKERS = min(5, cpu_count())  # CPUに応じて自動調整
CHUNK_SIZE = 2  # 1プロセスで処理するASIN数（適宜調整）


# --- Driver作成 ---
def create_driver():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-extensions")
    options.add_argument("user-agent=Mozilla/5.0 ...")
    service = Service(CHROME_DRIVER_PATH)
    return webdriver.Chrome(service=service, options=options)

def create_driver_non_headless():
    options = Options()
    # ヘッドレスモードなし
    options.add_argument("--disable-extensions")
    options.add_argument("user-agent=Mozilla/5.0 ...")
    service = Service(CHROME_DRIVER_PATH)
    return webdriver.Chrome(service=service, options=options)

# --- AmazonログインしてCookie取得 ---
def get_logged_in_cookies():
    from selenium.webdriver.common.by import By
    driver = create_driver_non_headless()
    try:
        driver.get(AMAZON_LOGIN_URL)
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "ap_email"))).send_keys(AMAZON_EMAIL)
        driver.find_element(By.ID, "continue").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "ap_password"))).send_keys(AMAZON_PASSWORD)
        driver.find_element(By.ID, "signInSubmit").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "nav-link-accountList-nav-line-1")))
        cookies = driver.get_cookies()
        return cookies
    finally:
        driver.quit()


# --- 単一ASIN情報取得 ---
def get_asin_info(driver, asin):
    try:
        print(f"🔍 アクセス中: https://www.amazon.co.jp/dp/{asin}")
        driver.get(f"https://www.amazon.co.jp/dp/{asin}")
        print(f"✅ {asin} ページ読み込み完了")
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//table[contains(@class, "prodDetTable")]'))
        )
        tables = driver.find_elements(By.XPATH, '//table[contains(@class, "prodDetTable")]')
        rows = driver.find_elements(By.XPATH, '//table[contains(@class, "prodDetTable")]//tr')
        height = width = length = weight = None
        
        dim_patterns = [
            r'(\d+(?:\.\d+)?)\s*[×xX＊*]\s*(\d+(?:\.\d+)?)\s*[×xX＊*]\s*(\d+(?:\.\d+)?)\s*(mm|cm)',
            r'奥行き\s*(\d+(?:\.\d+)?)\s*[×xX＊*]\s*幅\s*(\d+(?:\.\d+)?)\s*[×xX＊*]\s*高さ\s*(\d+(?:\.\d+)?)',
            r'幅\s*(\d+(?:\.\d+)?)\s*[×xX＊*]\s*奥行き\s*(\d+(?:\.\d+)?)\s*[×xX＊*]\s*高さ\s*(\d+(?:\.\d+)?)',
            r'(\d+(?:\.\d+)?)\s*奥行き\s*[×xX＊*]\s*(\d+(?:\.\d+)?)\s*幅\s*[×xX＊*]\s*(\d+(?:\.\d+)?)\s*高さ',
            r'(\d+(?:\.\d+)?)\s*長さ\s*[×xX＊*]\s*(\d+(?:\.\d+)?)\s*幅\s*[×xX＊*]\s*(\d+(?:\.\d+)?)\s*高さ'
        ]
        weight_pattern = r'(\d+(?:\.\d+)?)\s*(kg|キログラム|g|グラム|lb|lbs|ポンド)'

        for row in rows:
            try:
                th_elem = row.find_element(By.TAG_NAME, "th")
                td_elem = row.find_element(By.TAG_NAME, "td")
        
                th_text = th_elem.get_attribute("innerText").strip()
                td_text = td_elem.get_attribute("innerText").strip()
            except Exception as e:
                continue
            if "寸法" in th_text or "サイズ" in th_text:
                for pattern in dim_patterns:
                    m = re.search(pattern, td_text)
                    if m:
                        groups = list(m.groups())
                        if len(groups) == 4:
                            length, width, height, unit = groups
                        else:
                            length, width, height = groups
                            unit = "cm" if "cm" in td_text else "mm"
                        length = f"{length}{unit}"
                        width = f"{width}{unit}"
                        height = f"{height}{unit}"
                        break


                if not weight and "重量" in th_text:
                    m = re.search(weight_pattern, td_text, re.IGNORECASE)
                    if m:
                        val, unit = m.groups()
                        unit = unit.lower()
                        if unit in ['kg', 'キログラム']:
                            weight = f"{val}kg"
                        elif unit in ['g', 'グラム']:
                            weight = f"{val}g"
                        elif unit in ['lb', 'lbs', 'ポンド']:
                            weight = f"{float(val) * 0.453592:.2f}kg"
                        else:
                            weight = f"{val}({unit})"
                if all([height, width, length]) and weight:
                            break

            

        return {"ASIN": asin, "高さ": height, "幅": width, "奥行き（長さ）": length, "重量": weight}

    except Exception as e:
        return {"ASIN": asin, "高さ": None, "幅": None, "奥行き（長さ）": None, "重量": None}        

# --- チャンク単位の処理（プロセスごとに1 driver で複数ASIN） ---
def process_chunk(chunk_with_cookies):
    chunk, cookies = chunk_with_cookies
    driver = create_driver()
    try:
        driver.get("https://www.amazon.co.jp")
        for cookie in cookies:
            try:
                driver.add_cookie(cookie)
            except:
                continue
        driver.refresh()
        results = []
        for i, asin in enumerate(chunk):
            result = get_asin_info(driver, asin)
            results.append(result)
            if i % 5 == 0:
                time.sleep(random.uniform(2, 4))  # Amazon対策
        return results
    finally:
        driver.quit()


# --- メイン処理 ---
def main():
    df = pd.read_excel(INPUT_EXCEL, header=None)
    all_asins = df[0].dropna().astype(str).tolist()

    if os.path.exists(OUTPUT_EXCEL):
        done_asins = pd.read_excel(OUTPUT_EXCEL)['ASIN'].astype(str).tolist()
        asins = [a for a in all_asins if a not in done_asins]
    else:
        asins = all_asins

    print(f"処理ASIN数: {len(asins)}")

    chunks = [asins[i:i + CHUNK_SIZE] for i in range(0, len(asins), CHUNK_SIZE)]

    cookies = get_logged_in_cookies()

    pool = None
    try:
        pool = Pool(processes=NUM_WORKERS)
        for i, results in enumerate(pool.imap_unordered(process_chunk, [(chunk, cookies) for chunk in chunks])):
            df_chunk = pd.DataFrame(results)
            if not os.path.exists(OUTPUT_EXCEL):
                df_chunk.to_excel(OUTPUT_EXCEL, index=False)
            else:
                book = load_workbook(OUTPUT_EXCEL)
                sheet = book.active
                start_row = sheet.max_row

                with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_chunk.to_excel(writer, index=False, header=False, startrow=start_row)
            print(f"✅ チャンク {i+1}/{len(chunks)} 保存完了")
        pool.close()
        pool.join()
    except KeyboardInterrupt:
        print("\n🛑 中断要求を検出。全プロセスを強制終了します。")
        if pool is not None:
            pool.terminate()
            pool.join()

    print("処理完了")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n🛑 中断されました。全プロセスを終了します。")
